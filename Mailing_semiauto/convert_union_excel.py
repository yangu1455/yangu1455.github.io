import json
from openpyxl import load_workbook

workbook = load_workbook(
    "노동조합.xlsx"
)

sheet = workbook.active

numbers = []

# B열만 읽기
for row in sheet.iter_rows(
    min_col=2,
    max_col=2,
    values_only=True
):

    cell = row[0]

    if cell is None:
        continue

    number = str(cell)

    number = (
        number
        .replace("-", "")
        .replace(" ", "")
        .strip()
    )

    numbers.append(number)

with open(
    "union_numbers.json",
    "w",
    encoding="utf-8"
) as file:

    json.dump(
        numbers,
        file,
        ensure_ascii=False,
        indent=2
    )

print(
    f"총 {len(numbers)}개 저장 완료"
)