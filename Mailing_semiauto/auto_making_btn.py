from openpyxl import load_workbook

CHUNK_SIZE = 19

# 엑셀 읽기
wb = load_workbook("전화번호부.xlsx")
ws = wb.active

numbers = []

for row in ws.iter_rows(values_only=True):

    value = row[1]

    if value is None:
        continue

    digits = "".join(
        filter(str.isdigit, str(value))
    )

    if len(digits) == 8:
        digits = "010" + digits

    numbers.append(digits)

# 19개씩 분할
groups = [
    numbers[i:i + CHUNK_SIZE]
    for i in range(0, len(numbers), CHUNK_SIZE)
]

# 버튼 HTML 생성
buttons_html = ""

for idx, group in enumerate(groups, start=1):

    # LG U+
    lg_group = [
        f"{num}#"
        for num in group
    ]

    lg_recipients = ";".join(lg_group)

    # KT
    kt_group = [
        f"77*{num}"
        for num in group
    ]

    kt_recipients = ";".join(kt_group)

    buttons_html += f'''
    <a class="btn"
       href="sms:{lg_recipients}">
       문자 {idx} (LG)
    </a>

    <br><br>

    <a class="btn"
       href="sms:{kt_recipients}">
       문자 {idx} (KT)
    </a>

    <br><br>
'''

# 기존 html 읽기
with open(
    "mailing_semiauto.html",
    "r",
    encoding="utf-8"
) as f:

    html = f.read()

# 버튼 영역 교체
start = "<!-- BUTTONS_START -->"
end = "<!-- BUTTONS_END -->"

before = html.split(start)[0]
after = html.split(end)[1]

new_html = (
    before
    + start
    + "\n"
    + buttons_html
    + "\n"
    + end
    + after
)

# 저장
with open(
    "mailing_semiauto.html",
    "w",
    encoding="utf-8"
) as f:

    f.write(new_html)

print("완료")