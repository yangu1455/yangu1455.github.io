import json

from openpyxl import load_workbook

# 엑셀 읽기
wb = load_workbook("전화번호부.xlsx")

ws = wb.active

numbers = []

for row in ws.iter_rows(values_only=True):

    value = row[1]

    if value is None:
        continue

    # 숫자만 추출
    digits = "".join(
        filter(str.isdigit, str(value))
    )

    # 8자리면 앞에 010 붙이기
    if len(digits) == 8:
        digits = "010" + digits

    numbers.append(digits)

# 저장
with open(
    "numbers.json",
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        numbers,
        f,
        ensure_ascii=False,
        indent=2
    )

print("numbers.json 생성 완료")